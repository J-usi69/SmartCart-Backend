from io import BytesIO
from decimal import Decimal
from datetime import datetime
import logging

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

from openpyxl import Workbook
from openpyxl.styles import Font

from django.http import HttpResponse
from orders.models import Order, OrderItem

logger = logging.getLogger(__name__)


class ReportGenerator:

    
    def generate_pdf(self, title, headings, data, filename):
        try:
            buffer = BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=letter)
            elements = []
            

            styles = getSampleStyleSheet()
            title_style = styles['Heading1']
            

            elements.append(Paragraph(title, title_style))
            elements.append(Spacer(1, 20))
            

            table_data = [headings] + data
            t = Table(table_data)
            t.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            
            elements.append(t)
            doc.build(elements)
            
            buffer.seek(0)
            response = HttpResponse(buffer, content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
            return response
        except Exception as e:
            logger.exception(f"Error generando PDF: {str(e)}")
            return None
    
    def generate_excel(self, headings, data, filename):
        try:
            wb = Workbook()
            ws = wb.active
            

            for col_idx, header in enumerate(headings, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
            

            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            response = HttpResponse(
                buffer.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
            return response
        except Exception as e:
            logger.exception(f"Error generando Excel: {str(e)}")
            return None


class ClientReportGenerator(ReportGenerator):

    
    def get_report_data(self, client_id, start_date=None, end_date=None):
        orders = Order.objects.filter(client_id=client_id)
        
        if start_date:
            orders = orders.filter(created_at__gte=start_date)
        if end_date:
            end_of_day = datetime.combine(end_date.date(), datetime.max.time())
            orders = orders.filter(created_at__lte=end_of_day)
        
        orders_data = []
        items_data = []
        total_spent = Decimal('0')
        
        for order in orders:
            order_row = [
                order.id, 
                order.created_at.strftime("%d/%m/%Y %H:%M"), 
                order.status, 
                f"${order.total_price}"
            ]
            orders_data.append(order_row)
            total_spent += order.total_price
            
            items = OrderItem.objects.filter(order=order)
            for item in items:
                item_price = item.product.final_price
                subtotal = item_price * item.quantity
                items_data.append([
                    order.id,
                    item.product.name,
                    f"${item_price}",
                    item.quantity,
                    f"${subtotal}"
                ])
        
        return {
            'client_id': client_id,
            'orders': orders_data,
            'items': items_data,
            'total_spent': total_spent,
            'orders_count': len(orders_data)
        }
    
    def generate_pdf_report(self, client_id, start_date=None, end_date=None):
        data = self.get_report_data(client_id, start_date, end_date)
        
        title = f"Reporte de Cliente {client_id}"
        if not data['orders']:
            return HttpResponse("No hay datos para este cliente en el período seleccionado", status=404)
        

        orders_headings = ["ID", "Fecha", "Estado", "Total"]

        orders_data = data['orders'] + [["", "", "TOTAL", f"${data['total_spent']}"]]
        
        return self.generate_pdf(title, orders_headings, orders_data, f"cliente_{client_id}_reporte")
    
    def generate_excel_report(self, client_id, start_date=None, end_date=None):
        data = self.get_report_data(client_id, start_date, end_date)
        
        if not data['orders']:
            return HttpResponse("No hay datos para este cliente en el período seleccionado", status=404)
        
        wb = Workbook()
        

        ws_orders = wb.active
        ws_orders.title = "Órdenes"
        

        headers = ["ID", "Fecha", "Estado", "Total"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws_orders.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True)
        

        for row_idx, order_data in enumerate(data['orders'], 2):
            for col_idx, value in enumerate(order_data, 1):
                ws_orders.cell(row=row_idx, column=col_idx, value=value)
        

        total_row = row_idx + 1
        ws_orders.cell(row=total_row, column=3, value="TOTAL")
        ws_orders.cell(row=total_row, column=4, value=f"${data['total_spent']}")
        

        if data['items']:
            ws_items = wb.create_sheet(title="Detalles")
            

            headers = ["Orden ID", "Producto", "Precio", "Cantidad", "Subtotal"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws_items.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True)
            

            for row_idx, item_data in enumerate(data['items'], 2):
                for col_idx, value in enumerate(item_data, 1):
                    ws_items.cell(row=row_idx, column=col_idx, value=value)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename = f"cliente_{client_id}_reporte.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        # Forma correcta de establecer el Content-Disposition
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


class TopProductsReportGenerator(ReportGenerator):

    
    def get_report_data(self, start_date=None, end_date=None, limit=10):
        from django.db.models import Sum, F, ExpressionWrapper, DecimalField
        
        query = OrderItem.objects.values('product__id', 'product__name')
        
        if start_date:
            query = query.filter(order__created_at__gte=start_date)
        if end_date:
            end_of_day = datetime.combine(end_date.date(), datetime.max.time())
            query = query.filter(order__created_at__lte=end_of_day)
        
        top_products = query.annotate(
            total_sold=Sum('quantity'),
            total_revenue=Sum(
                ExpressionWrapper(
                    F('product__price') * F('quantity'),
                    output_field=DecimalField()
                )
            )
        ).order_by('-total_sold')[:limit]
        
        products_data = []
        for product in top_products:
            products_data.append([
                product['product__id'],
                product['product__name'],
                product['total_sold'],
                f"${product['total_revenue']}"
            ])
        
        return products_data
    
    def generate_pdf_report(self, start_date=None, end_date=None, limit=10):
        data = self.get_report_data(start_date, end_date, limit)
        
        if not data:
            return HttpResponse("No hay datos para el período seleccionado", status=404)
        
        title = "Productos Más Vendidos"
        headings = ["ID", "Producto", "Cantidad Vendida", "Ingresos"]
        
        return self.generate_pdf(title, headings, data, "productos_mas_vendidos")
    
    def generate_excel_report(self, start_date=None, end_date=None, limit=10):
        data = self.get_report_data(start_date, end_date, limit)
        
        if not data:
            return HttpResponse("No hay datos para el período seleccionado", status=404)
        
        headings = ["ID", "Producto", "Cantidad Vendida", "Ingresos"]
        
        return self.generate_excel(headings, data, "productos_mas_vendidos")

